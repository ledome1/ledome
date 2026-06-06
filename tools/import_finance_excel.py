from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = Path(r"C:\Users\Admin\OneDrive\Desktop\LE DOME - QUAN LY THU CHI 2026 (1).xlsx")
DEFAULT_OUTPUT = ROOT / "data" / "runtime-finance.json"

PROJECT_GROUP = "DỰ ÁN"
OPERATION_GROUP = "VẬN HÀNH LE DOME"
FINANCE_GROUP = "TÀI CHÍNH"
CONTRACT_CATEGORIES = ["HỢP ĐỒNG THIẾT KẾ", "HỢP ĐỒNG THI CÔNG"]
FINANCE_TOPICS = {"VAY", "NGÂN HÀNG", "LÃI VAY", "PHÍ NGÂN HÀNG"}


def clean(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def clean_upper(value) -> str:
    return clean(value).upper()


def number(value) -> int:
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    text = re.sub(r"[^\d.-]", "", str(value))
    return int(round(float(text))) if text else 0


def iso_date(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text[:10]


def unique(values) -> list[str]:
    seen = set()
    out = []
    for value in values:
        item = clean(value)
        if not item:
            continue
        key = item.casefold()
        if key not in seen:
            seen.add(key)
            out.append(item)
    return out


def append_note(note: str, extra: str) -> str:
    note = clean(note)
    return f"{note} | {extra}" if note else extra


def clean_warning(value, profit: int) -> str:
    text = clean(value)
    text = re.sub(r"^[✅⚠️\s]+", "", text).strip()
    if text:
        return text
    return "Cần rà soát" if profit < 0 else "Ổn"


def customer_name(value) -> str:
    text = clean(value)
    return "Chưa cập nhật CĐT" if not text or set(text) <= {"?"} else text


def get_sheet(wb, name: str, index: int):
    return wb[name] if name in wb.sheetnames else wb.worksheets[index]


def read_partner_names(kind: str) -> list[str]:
    path = ROOT / "data" / "runtime-partners.json"
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    return [row[1] for row in data.get(kind, []) if len(row) > 1 and clean(row[1])]


def read_projects(ws) -> list[list]:
    projects = []
    for row in ws.iter_rows(min_row=4, max_col=19, values_only=True):
        if row[0] is None:
            break
        old_revenue = number(row[10])
        old_cost = number(row[11])
        actual_revenue = number(row[12])
        actual_cost = number(row[13])
        profit = number(row[14])
        projects.append([
            clean_upper(row[0]),
            clean(row[1]),
            customer_name(row[2]),
            clean(row[3]),
            clean(row[4]),
            clean(row[5]),
            old_revenue,
            old_cost,
            old_revenue + actual_revenue,
            old_cost + actual_cost,
            profit,
            number(row[16]),
            number(row[17]),
            clean_warning(row[18], profit),
        ])
    return projects


def read_catalog_categories(ws) -> list[str]:
    values = []
    for row in ws.iter_rows(min_row=3, max_col=7, values_only=True):
        if clean_upper(row[1]) == PROJECT_GROUP and clean(row[3]):
            values.append(clean_upper(row[3]))
    return unique([*CONTRACT_CATEGORIES, *values])


def read_transactions(ws, projects: list[list]) -> tuple[list[list], list[dict]]:
    project_by_code = {row[0]: row for row in projects}
    project_codes = set(project_by_code)
    transactions = []
    reclassified = []

    for excel_row, row in enumerate(ws.iter_rows(min_row=5, max_col=13, values_only=True), start=5):
        tx_type = clean(row[1])
        amount = number(row[7])
        if not row[0] or tx_type not in {"Thu", "Chi"} or amount == 0:
            continue

        original_group = clean_upper(row[2])
        topic = clean_upper(row[3])
        category = clean_upper(row[4])
        partner = clean(row[5])
        note = clean(row[12])

        if original_group == "LE DOME":
            group = OPERATION_GROUP
        elif original_group == PROJECT_GROUP and topic in project_codes:
            group = PROJECT_GROUP
        elif topic in FINANCE_TOPICS or category == "NGÂN HÀNG":
            group = FINANCE_GROUP
        elif original_group == PROJECT_GROUP:
            group = FINANCE_GROUP
            note = append_note(note, f"Nguồn Excel dòng {excel_row}: DỰ ÁN/{topic or 'trống'} thiếu mã dự án")
            reclassified.append({"row": excel_row, "topic": topic, "category": category, "amount": amount})
        else:
            group = FINANCE_GROUP

        if group == PROJECT_GROUP:
            if tx_type == "Thu" and (not partner or clean_upper(partner) == "LE DOME"):
                partner = project_by_code.get(topic, ["", "", "Chưa cập nhật CĐT"])[2]
            elif tx_type == "Chi" and not partner:
                partner = "KHÁC"
        elif group == OPERATION_GROUP:
            topic = topic or "VẬN HÀNH"
            category = category or "KHÁC"
            partner = partner or "LE DOME"
        else:
            topic = topic or "KHÁC"
            category = category or "KHÁC"
            partner = partner or "LE DOME"

        transactions.append([
            iso_date(row[0]),
            tx_type,
            group,
            topic,
            category,
            partner,
            clean(row[6]),
            amount,
            number(row[8]),
            clean_upper(row[9]),
            note,
        ])

    return transactions, reclassified


def build_budgets(transactions: list[list], projects: list[list]) -> list[list]:
    project_name = {row[0]: row[1] for row in projects}
    buckets = {}
    for row in transactions:
        if row[1] != "Chi" or row[2] != PROJECT_GROUP:
            continue
        key = (row[3], row[4] or "KHÁC", row[5] or "KHÁC")
        current = buckets.setdefault(key, {"spent": 0, "extra": 0})
        current["spent"] += number(row[7])
        if row[9] == "PHÁT SINH":
            current["extra"] += number(row[7])

    budgets = []
    for (code, category, partner), values in sorted(buckets.items()):
        spent = values["spent"]
        extra = values["extra"]
        budgets.append([
            code,
            project_name.get(code, code),
            category,
            partner,
            spent,
            spent - extra,
            extra,
            spent,
            0,
            100,
            "Có phát sinh" if extra else "Đã chi",
        ])
    return budgets


def build_options(transactions: list[list], projects: list[list], catalog_categories: list[str]) -> dict:
    project_codes = [row[0] for row in projects]
    project_customers = [row[2] for row in projects if row[2] != "Chưa cập nhật CĐT"]
    project_income_partners = [
        row[5] for row in transactions
        if row[2] == PROJECT_GROUP and row[1] == "Thu"
    ]
    project_expense_partners = [
        row[5] for row in transactions
        if row[2] == PROJECT_GROUP and row[1] == "Chi"
    ]
    operation_rows = [row for row in transactions if row[2] == OPERATION_GROUP]
    finance_rows = [row for row in transactions if row[2] == FINANCE_GROUP]

    customers = unique([*project_customers, *read_partner_names("customers"), *project_income_partners])
    contractors = unique([*read_partner_names("contractors"), *project_expense_partners])
    suppliers = unique([*read_partner_names("suppliers"), *project_expense_partners])

    return {
        "groups": [PROJECT_GROUP, OPERATION_GROUP, FINANCE_GROUP],
        "projectCodes": project_codes,
        "projectCategories": unique([
            *CONTRACT_CATEGORIES,
            *catalog_categories,
            *(row[4] for row in transactions if row[2] == PROJECT_GROUP),
        ]),
        "projectIncomePartners": customers,
        "projectExpensePartners": unique([*contractors, *suppliers]),
        "operationTopics": unique(row[3] for row in operation_rows),
        "operationCategories": unique(row[4] for row in operation_rows),
        "operationPartners": unique(row[5] for row in operation_rows),
        "financeTopics": unique(row[3] for row in finance_rows),
        "financeCategories": unique(row[4] for row in finance_rows),
        "financePartners": unique(row[5] for row in finance_rows),
    }


def build_finance_data(workbook_path: Path) -> dict:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    transactions_ws = get_sheet(wb, "Giao dịch", 2)
    projects_ws = get_sheet(wb, "Dự án", 3)
    catalog_ws = get_sheet(wb, "Danh mục", 9)

    projects = read_projects(projects_ws)
    transactions, reclassified = read_transactions(transactions_ws, projects)
    catalog_categories = read_catalog_categories(catalog_ws)

    return {
        "transactions": transactions,
        "projects": projects,
        "budgets": build_budgets(transactions, projects),
        "debts": [],
        "options": build_options(transactions, projects, catalog_categories),
        "meta": {
            "sourceWorkbook": str(workbook_path),
            "importedAt": datetime.now().isoformat(timespec="seconds"),
            "transactionCount": len(transactions),
            "projectCount": len(projects),
            "ignoredSheets": ["Danh mục", "Công nợ"],
            "reclassifiedRows": reclassified,
        },
    }


def main() -> int:
    workbook_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT
    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 1

    data = build_finance_data(workbook_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {output_path}")
    print(f"Transactions: {len(data['transactions'])}, projects: {len(data['projects'])}, budgets: {len(data['budgets'])}")
    if data["meta"]["reclassifiedRows"]:
        print(f"Reclassified rows: {len(data['meta']['reclassifiedRows'])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
